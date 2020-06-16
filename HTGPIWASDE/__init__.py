import logging
import azure.functions as func
from azure.storage.blob import BlockBlobService, PublicAccess
import requests
import urllib.request as ulib
from io import BytesIO
import openpyxl as xl
import xlrd
import time
import pandas as pd
import pyodbc
#import re
from bs4 import BeautifulSoup
import datetime

def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')
    name = req.params.get('name')
    if not name:
        try:
            req_body = req.get_json()
        except ValueError:
            pass
        else:
            name = req_body.get('name')

    if name:
        respuesta = upload_azure()
        #return func.HttpResponse(f"Hello {name}! "+respuesta)
        return func.HttpResponse(respuesta)
    else:
        return func.HttpResponse(
             "Debe ingresar el name",
             status_code=400
        )

def upload_azure():
    # Create the BlockBlockService that is used to call the Blob service for the storage account
    block_blob_service = BlockBlobService(account_name='gpistore', account_key='zfKM5R0PuPwR0F+pPsgs5BW/AQjAxv5fwKojoP2W38II++qfT6e+axFrRAcTOmKi/8U0tyJbrB2A3XCd7W7o6A==')

    # Create a container called 'quickstartblobs'.
    container_name ='gpistore'
    block_blob_service.create_container(container_name)

    # Set the permission so the blobs are public.
    block_blob_service.set_container_acl(container_name, public_access=PublicAccess.Container)
    
    #Verificar si el archivo ya existe en WASDE
    url = "https://usda.library.cornell.edu/concern/publications/3t945q76s?locale=en"      
    if exist_fileurl(url)==False:
        return 'False'

    #Obtenemos el ultimo archivo publicado por Wasde
    req = requests.get(url)
    html = BeautifulSoup(req.content, "html.parser")
    entradas= html.find_all('tr', {'class' : 'release attributes row'})
    
    #FJSOLIS
    #texto = entradas[0].find('td',{'class' : 'attribute date_uploaded'}).text
    #link = entradas[0].find('td',{'class' : 'file_set'}).find('a',{'data-label' : re.compile(r'[A-Za-z]*[0-9]*\.xls$')}).attrs['href']
    #daterelease = entradas[0].find('td',{'class' : 'file_set'}).find('a',{'data-label' : re.compile(r'[A-Za-z]*[0-9]*\.xls$')}).attrs['data-release-date']
    #daterelease = daterelease[:10]

    #ORIGINAL
    texto = entradas[0].find('td',{'class' : 'attribute date_uploaded'}).text
    link = entradas[0].find('td',{'class' : 'file_set'}).find('a',{'data-label' : 'latest.xls'}).attrs['href']
    daterelease = entradas[0].find('td',{'class' : 'file_set'}).find('a',{'data-label' : 'latest.xls'}).attrs['data-release-date']
    daterelease = daterelease[:10]
    
    #Verificamos si el daterealese del archivo ya fue cargado en SQL, si fue asi se termina la ejecucion
    if (exist_daterelease(daterelease)):
        return 'False'
    
    #Verificamos que el mes no exista, si existe lo eliminamos antes de cargar esta nueva version
    #Esto se hace porque pueden haber varias versiones en un mismo mes y solo necesitamos la ultima
    if (exist_monthyear(daterelease)):
        delete_wasde(daterelease)

    fname = daterelease.replace('-','')
    name = 'WASDE_'+daterelease+'.xls'
    req = ulib.urlopen(link)
    req_read = req.read()
    file_download = BytesIO(req_read)
    
    # Carga del archivo completo de WASDE al azure blob storage
    block_blob_service.create_blob_from_stream(container_name,name,file_download)

    #Extraccion de datos
    name = 'WASDE_STOCK_TO_USE.csv'
    file_download = extraer_datos(req_read,fname,daterelease)
    block_blob_service.create_blob_from_text(container_name,name,file_download)
    return 'True'

def exist_daterelease(daterelease):
    server = 'grainpredictive.database.windows.net'
    database = 'gpi'
    username = 'gpi'
    password = 'Cmi@2019$A'
    driver= '{ODBC Driver 17 for SQL Server}'
    cnxn = pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    sql = "SELECT * FROM [ST_WASDE].[STOCKS_TO_USE] where [DATERELEASE ] = '"+daterelease+"'"
    n = cursor.execute(sql)
    if(n.rowcount==0):
        return False
    else:
        return True

def exist_monthyear(daterelease):
    server = 'grainpredictive.database.windows.net'
    database = 'gpi'
    username = 'gpi'
    password = 'Cmi@2019$A'
    driver= '{ODBC Driver 17 for SQL Server}'
    cnxn = pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    sql = "SELECT * FROM [ST_WASDE].[STOCKS_TO_USE] where left(DATERELEASE,7)  = '"+daterelease[:7]+"'"
    n = cursor.execute(sql)
    if(n.rowcount==0):
        return False
    else:
        return True

def delete_wasde(daterelease):
    server = 'grainpredictive.database.windows.net'
    database = 'gpi'
    username = 'gpi'
    password = 'Cmi@2019$A'
    driver= '{ODBC Driver 17 for SQL Server}'
    cnxn = pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    sql = "delete from [ST_WASDE].[STOCKS_TO_USE] where left(DATERELEASE,7) = '"+daterelease[:7]+"' "
    cursor.execute(sql)
    cursor.commit()

def extraer_datos(req_read,fname,daterelease):
    headers=['Origen','Archivo','Wasde','Datos','Commoditie','Medida','DateN','HarvestDate','Tipo','Grupo','Geography','Orden','Mes','Beginning stocks','Production','Imports','Domestic Feed','Domestic total','Exports','Ending stocks','Total Use','Stocks to Use']
    
    #Pagina 11
    datos = 'U.S. Wheat by Class: Supply and Use'
    pagina = 'Page 11'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p11(data,fname,datos)                    
    df11 = pd.DataFrame(processed_data,columns=headers)

    #Pagina 18
    datos = 'World Wheat Supply and Use'
    pagina = 'Page 18'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p18(data,fname,datos)                    
    df18 = pd.DataFrame(processed_data,columns=headers)

    #Pagina 19
    datos = 'World Wheat Supply and Use'
    pagina = 'Page 19'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p19(data,fname,datos)                    
    df19 = pd.DataFrame(processed_data,columns=headers)
    
    #Pagina 22
    datos = 'World Corn Supply and Use'
    pagina = 'Page 22'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p22(data,fname,datos)                    
    df22 = pd.DataFrame(processed_data,columns=headers)

    #Pagina 23
    datos = 'World Corn Supply and Use'
    pagina = 'Page 23'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p23(data,fname,datos)                    
    df23 = pd.DataFrame(processed_data,columns=headers)

    #Pagina 28
    datos = 'World Soybean Supply & Use'
    pagina = 'Page 28'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p28(data,fname,datos)                    
    df28 = pd.DataFrame(processed_data,columns=headers)

    #Pagina 29
    datos = 'World Soybean Meal Supply & Use'
    pagina = 'Page 29'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p29(data,fname,datos)                         
    df29 = pd.DataFrame(processed_data,columns=headers)

    #Pagina 30
    datos = 'World Soybean Oil Supply & Use'
    pagina = 'Page 30'
    data=read_xls_with_sheetname(req_read,pagina) 
    processed_data=process_data_p30(data,fname,datos)                    
    df30 = pd.DataFrame(processed_data,columns=headers)

    frames = [df11,df18,df19,df22,df23,df28,df29,df30]##df11,df18,df19,df22,df23,df28,df29
    df = pd.concat(frames)
    df['actualizacion'] = datetime.datetime.now()
    df['DATERELEASE'] = daterelease

    combined_csv = df.to_csv(header=True,index=False,encoding='utf-8-sig')
    return combined_csv

def read_xls_with_sheetname(filename,sheet_name):
    '''read specific sheet by name in excel(xls) file and returns a list of lists '''
    data=[]
    book_xls = xlrd.open_workbook(file_contents=filename, formatting_info=True, ragged_rows=True)

    book_xlsx = xl.workbook.Workbook()
    sheet_names = book_xls.sheet_names()
    for sheet_index in range(len(sheet_names)):
        if sheet_name !=sheet_names[sheet_index]:
            continue
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
        for crange in sheet_xls.merged_cells:
            rlo, rhi, clo, chi = crange
            sheet_xlsx.merge_cells(start_row=rlo + 1, end_row=rhi,
            start_column=clo + 1, end_column=chi,)

        def _get_xlrd_cell_value(cell):
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                datetime_tup = xlrd.xldate_as_tuple(value,0)    
                if datetime_tup[0:3] == (0, 0, 0):   # time format without date
                    value = datetime.time(*datetime_tup[3:])
                else:
                    value = datetime.datetime(*datetime_tup)
            return value

        for row in range(sheet_xls.nrows):
            new_row=[]
            try:
                for col in (_get_xlrd_cell_value(cell) for cell in sheet_xls.row_slice(row, end_colx=sheet_xls.row_len(row))):
                    new_row.append(col)
            except:
                continue
            if len(new_row) != 0:
                data.append(new_row)
  
    end = time.time()   
    return data

#FJSOLIS - INICIO
def process_data_p11(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]    
    est_data=[]
    proj_data=[]

    header_index=0
    empty_index=0
    date_index=0
    hard_red_winter_index = 0
    hard_red_spring_index = 0
    soft_red_winter_index = 0
    white_index = 0
    durum_index = 0
    total_index = 0

    US_WbyClass = False

    medida = ''
    wasde = ''
    
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row) and wasde == '':
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell

        if 'U.S. Wheat by Class: Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'U.S. Wheat by Class: Supply and Use' in cell:
                    US_WbyClass = True

        if US_WbyClass and 'Million' in str(row):
            for index,cell in enumerate(row):
                if 'Million' in cell:
                    medida = cell
                if 'Bushels' in cell:
                    medida += ' ' + cell
                if medida == 'Million Bushels':
                    break

        if US_WbyClass and 'Year beginning June 1' in row:
            header_index=i
            for index,cell in enumerate(row):
                if not cell.find('Year') == -1:
                    empty_index=index
                    date_index=index
                if 'Hard Red\nWinter' in cell or 'Hard\nWinter' in cell:
                    hard_red_winter_index=index
                if 'Hard Red\nSpring' in cell or 'Hard\nSpring' in cell:
                    hard_red_spring_index=index
                if 'Soft Red\nWinter' in cell or 'Soft\nRed' in cell:
                    soft_red_winter_index=index
                if 'White' in cell:
                    white_index=index
                if "Durum" in cell:
                    durum_index=index
                if 'Total' in cell:
                    total_index=index
    date=''

    #extracting data from cols
    colname_ant = ''
    colname = 'Commoditie'
    tipo = 'Header'
    harvest_date = []
    for row in data[header_index:]:
        new_row=[]
        
        if (row[empty_index]=='' and row[empty_index + 1]=='') or 'Note:  Totals may not add due to rounding.' in row[date_index]:
            continue
        if not row[empty_index + 1]=='':
            colname = row[empty_index + 1]
        else:
            if not row[empty_index + 2]=='':
                colname = row[empty_index + 2]
                empty_index+=1
        if colname == colname_ant:
            continue

        new_row.append(colname)
        new_row.append(row[hard_red_winter_index])
        new_row.append(row[hard_red_spring_index])
        new_row.append(row[soft_red_winter_index])
        new_row.append(row[white_index])
        new_row.append(row[durum_index])
        new_row.append(row[total_index])
        
        if not row[date_index]=='':
            if row[date_index].replace('(Est.) ','').replace('/','').replace(' ','').isdigit():
                harvest_date.append(row[date_index].replace('(Est.) ',''))
                tipo = 'EST.'
            if row[date_index].replace('(Proj.) ','').replace('/','').replace(' ','').isdigit():
                harvest_date.append(row[date_index].replace('(Proj.) ',''))
                tipo = 'PROJ.'
        if tipo=='EST.':
            est_data.append(new_row)
        if tipo=='PROJ.':
            proj_data.append(new_row)
        if tipo=='Header':
            est_data.append(new_row)
            est_data.append(new_row)
            proj_data.append(new_row)
            proj_data.append(new_row)
        colname_ant = colname

    np_est_data = np.array(est_data)
    np_proj_data = np.array(proj_data)

    df = pd.DataFrame(data=np_est_data[1:,1:],index=np_est_data[1:,0],columns=np_est_data[0,1:]).T
    
    df.insert(1, 'DateN', 2)
    df.insert(2, 'HarvestDate', harvest_date[0])
    df.insert(3, 'Tipo','EST.')
    df.insert(4, 'Grupo','RESUMEN')
    df.insert(5, 'Geography','UNITED STATES')
    df.insert(6, 'Orden',2)
    df.insert(7, 'Mes', obtener_mes(mid(archivo,4,2)).upper())

    if len(harvest_date) == 2:
        df_2 = pd.DataFrame(data=np_proj_data[1:,1:],index=np_proj_data[1:,0],columns=np_proj_data[0,1:]).T
        df_2.insert(1, 'DateN', 3)
        df_2.insert(2, 'HarvestDate', harvest_date[1])
        df_2.insert(3, 'Tipo','PROJ.')
        df_2.insert(4, 'Grupo','RESUMEN')
        df_2.insert(5, 'Geography','UNITED STATES')
        df_2.insert(6, 'Orden',2)
        df_2.insert(7, 'Mes', obtener_mes(mid(archivo,4,2)).upper())
        df = df.append(df_2)
    
    #Preparar el dataframe:
    df.insert(0, 'Origen', 'XLS')
    df.insert(1, 'Archivo', archivo)
    df.insert(2, 'Wasde', wasde.replace(' ','').strip().upper())
    df.insert(3, 'Datos', datos.upper())
    df.insert(5, 'Medida', medida.upper())

    #Calculos:
    #actualizar import_index
    df['  Supply, Total 3/'] = pd.to_numeric(df['  Supply, Total 3/'])-pd.to_numeric(df['Production'])-pd.to_numeric(df['Beginning Stocks']) 
    #copiar col al final
    df['Total_Use'] = df['  Use, Total'] 
    #rellenar valores vacios
    df.insert(16, 'Domestic_Feed', '')
    #eliminar extra
    del df['  Use, Total']
    #stocks to use
    df['Stocks_to_Use'] = (pd.to_numeric(df['Ending Stocks, Total'])/pd.to_numeric(df['Total_Use'])) * 100

    #Limpiar los datos.
    df.columns = df.columns.str.strip()
    df = df.replace('\n', ' ', regex=True)
    df_trimmed = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    #df a []
    new_data = df.to_numpy().tolist()
    return new_data

def process_data_p18(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]    
    header_index=0
    date_index=0
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Wheat Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Wheat Supply and Use' in cell:
                    commoditie = 'Wheat'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell.replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if 'Domestic\nFeed' in cell:
                    domestic_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''

    #extracting data from cols
    n = 2
    for row in data[header_index:]:
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            if n == 2:
                n = 1
            else:
                n = 2
            continue
        if row[date_index]=='' or '1/ Aggregate of local marketing years' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'
        new_row.append('XLS')
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())  
        new_row.append(n)
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')            
        new_row.append(grupo.upper())
        new_row.append(replace_string(row[date_index].upper().strip()))
        new_row.append('2')#Orden columnas vacias porque solo se usan en pag 19
        new_row.append(obtener_mes(mid(archivo,4,2)).upper())#Mes columnas vacias porque solo se usan en pag 19        
        new_row.append(row[begin_stock_index])
        new_row.append(row[production_index])
        new_row.append(row[import_index])
        new_row.append(row[domestic_index])
        new_row.append(row[domestic_index_2])
        new_row.append(row[exports_index])
        new_row.append(row[ending_stock_index])
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
    return new_data

def process_data_p19(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]
    header_index=0
    date_index=0
    report_date_index=0   
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Wheat Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Wheat Supply and Use' in cell:
                    commoditie = 'Wheat'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell                    
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell=='':
                    continue
                if  cell.split()[0].replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                    report_date_index = begin_stock_index-1
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if 'Domestic\nFeed' in cell:
                    domestic_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''
    #extracting data from cols
    geoant = ""
    for index,row in enumerate(data[header_index:]):
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            continue
        if (row[report_date_index]=='' and row[date_index] =='') or '1/ Aggregate of local marketing years' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'                    
        
        new_row.append('XLS')        
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())        
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())                         
        new_row.append('3')
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')
        new_row.append(grupo.upper())                    
        if row[date_index].strip() == '' and row[report_date_index].strip() != '':
            row[date_index] = data[header_index:][index - 1][date_index]
        geo = replace_string(row[date_index].upper().strip())
        if geo == geoant:
            orden = 2
        else:
            orden = 1
        new_row.append(geo)  #add geographic
        new_row.append(orden)
        new_row.append(row[report_date_index].upper())
        new_row.append(convert_value(row[begin_stock_index]))
        new_row.append(convert_value(row[production_index]))
        new_row.append(convert_value(row[import_index]))
        new_row.append(convert_value(row[domestic_index]))
        new_row.append(convert_value(row[domestic_index_2]))
        new_row.append(convert_value(row[exports_index]))
        new_row.append(convert_value(row[ending_stock_index]))
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
        geoant = geo
    return new_data

def process_data_p22(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]    
    header_index=0
    date_index=0
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Corn Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Corn Supply and Use' in cell:
                    commoditie = 'Corn'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell.replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if 'Domestic\nFeed' in cell:
                    domestic_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''

    #extracting data from cols
    n = 2
    for row in data[header_index:]:
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            if n == 2:
                n = 1
            else:
                n = 2
            continue
        if row[date_index]=='' or '1/ Aggregate of local marketing years' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'
        new_row.append('XLS')
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())  
        new_row.append(n)
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')            
        new_row.append(grupo.upper())
        new_row.append(replace_string(row[date_index].upper().strip()))
        new_row.append('2')#Orden columnas vacias porque solo se usan en pag 23
        new_row.append(obtener_mes(mid(archivo,4,2)).upper())#Mes columnas vacias porque solo se usan en pag 23        
        new_row.append(row[begin_stock_index])
        new_row.append(row[production_index])
        new_row.append(row[import_index])
        new_row.append(row[domestic_index])
        new_row.append(row[domestic_index_2])
        new_row.append(row[exports_index])
        new_row.append(row[ending_stock_index])
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
    return new_data

def process_data_p23(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]
    header_index=0
    date_index=0
    report_date_index=0   
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Corn Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Corn Supply and Use' in cell:
                    commoditie = 'Corn'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell                    
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell=='':
                    continue
                if  cell.split()[0].replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                    report_date_index = begin_stock_index-1
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if 'Domestic\nFeed' in cell:
                    domestic_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''
    #extracting data from cols
    geoant = ""
    for index,row in enumerate(data[header_index:]):
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            continue
        if (row[report_date_index]=='' and row[date_index] =='') or '1/ Aggregate of local marketing years' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'                    
        
        new_row.append('XLS')        
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())        
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())                         
        new_row.append('3')
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')
        new_row.append(grupo.upper())                    
        if row[date_index].strip() == '' and row[report_date_index].strip() != '':
            row[date_index] = data[header_index:][index - 1][date_index]
        geo = replace_string(row[date_index].upper().strip())
        if geo == geoant:
            orden = 2
        else:
            orden = 1
        new_row.append(geo)  #add geographic
        new_row.append(orden)
        new_row.append(row[report_date_index].upper())
        new_row.append(convert_value(row[begin_stock_index]))
        new_row.append(convert_value(row[production_index]))
        new_row.append(convert_value(row[import_index]))
        new_row.append(convert_value(row[domestic_index]))
        new_row.append(convert_value(row[domestic_index_2]))
        new_row.append(convert_value(row[exports_index]))
        new_row.append(convert_value(row[ending_stock_index]))
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
        geoant = geo
    return new_data

def process_data_p28(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]    
    header_index=0
    date_index=0
    report_date_index=0   
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Soybean Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Soybean Supply and Use' in cell:
                    commoditie = 'Soybean'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell=='':
                    continue
                if cell.replace('/','').isdigit() or cell.split()[0].replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                    report_date_index = begin_stock_index-1
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if 'Domestic\nCrush' in cell:
                    domestic_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''
    #extracting data from cols
    n = 0
    geoant = ""
    for index,row in enumerate(data[header_index:]):
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            if n == 0:
                n = 1
            elif n == 1:
                n = 2
            else:
                n = 3
            continue
        if (row[report_date_index]=='' and row[date_index] =='') or '1/ Data based on local marketing years except' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'
        new_row.append('XLS')
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())  
        new_row.append(n)
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')            
        new_row.append(grupo.upper())
        if row[date_index].strip() == '' and row[report_date_index].strip() != '':
            row[date_index] = data[header_index:][index - 1][date_index]
        geo = replace_string(row[date_index].upper().strip())
        if geo == geoant:
            orden = 2
        else:
            if n == 1 or n == 2:
                orden = 2
            elif n == 3:
                orden = 1
        new_row.append(geo)  #add geographic
        new_row.append(orden)
        if  row[report_date_index].upper() == '':
            new_row.append(obtener_mes(mid(archivo,4,2)).upper())#Mes columnas vacias porque solo se usan en pag 19        
        else:
            new_row.append(row[report_date_index].upper())
        new_row.append(convert_value(row[begin_stock_index]))
        new_row.append(convert_value(row[production_index]))
        new_row.append(convert_value(row[import_index]))
        new_row.append(convert_value(row[domestic_index]))
        new_row.append(convert_value(row[domestic_index_2]))
        new_row.append(convert_value(row[exports_index]))
        new_row.append(convert_value(row[ending_stock_index]))
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
        geoant = geo
    return new_data

def process_data_p29(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]    
    header_index=0
    date_index=0
    report_date_index=0   
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Soybean Meal Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Soybean Meal Supply and Use' in cell:
                    commoditie = 'Soybean Meal'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell=='':
                    continue
                if cell.replace('/','').isdigit() or cell.split()[0].replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                    report_date_index = begin_stock_index-1
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''
    #extracting data from cols
    n = 0
    geoant = ""
    for index,row in enumerate(data[header_index:]):
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            if n == 0:
                n = 1
            elif n == 1:
                n = 2
            else:
                n = 3
            continue
        if (row[report_date_index]=='' and row[date_index] =='') or '1/ Data based on local marketing years except' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'
        new_row.append('XLS')
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())  
        new_row.append(n)
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')            
        new_row.append(grupo.upper())
        if row[date_index].strip() == '' and row[report_date_index].strip() != '':
            row[date_index] = data[header_index:][index - 1][date_index]
        geo = replace_string(row[date_index].upper().strip())
        if geo == geoant:
            orden = 2
        else:
            if n == 1 or n == 2:
                orden = 2
            elif n == 3:
                orden = 1
        new_row.append(geo)  #add geographic
        new_row.append(orden)
        if  row[report_date_index].upper() == '':
            new_row.append(obtener_mes(mid(archivo,4,2)).upper())#Mes columnas vacias porque solo se usan en pag 19        
        else:
            new_row.append(row[report_date_index].upper())
        new_row.append(convert_value(row[begin_stock_index]))
        new_row.append(convert_value(row[production_index]))
        new_row.append(convert_value(row[import_index]))
        new_row.append('')
        new_row.append(convert_value(row[domestic_index_2]))
        new_row.append(convert_value(row[exports_index]))
        new_row.append(convert_value(row[ending_stock_index]))
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
        geoant = geo
    return new_data

def process_data_p30(data,archivo,datos):
    '''Extract data from Table based on provided instructions'''
    new_data=[]    
    header_index=0
    date_index=0
    report_date_index=0   
    begin_stock_index=0
    production_index=0
    import_index=0
    domestic_index_2=0
    exports_index=0
    ending_stock_index=0
    #Specify the cols index for table header(have smart detection for cols)
    for i,row in enumerate(data):
        if 'WASDE' in str(row):
            for index,cell in enumerate(row):
                if 'WASDE' in cell:
                    wasde = cell
        if 'World Soybean Oil Supply and Use' in str(row):
            for index,cell in enumerate(row):
                if 'World Soybean Oil Supply and Use' in cell:
                    commoditie = 'Soybean Oil'
        if 'Million Metric Tons' in str(row):
            for index,cell in enumerate(row):
                if 'Million Metric Tons' in cell:
                    medida = cell
        if 'Beginning\nStocks' in row:
            header_index=i
            for index,cell in enumerate(row):
                if cell=='':
                    continue
                if cell.replace('/','').isdigit() or cell.split()[0].replace('/','').isdigit():
                    date_index=index
                if cell=='Beginning\nStocks':
                    begin_stock_index=index
                    report_date_index = begin_stock_index-1
                if cell=='Production':
                    production_index=index
                if cell=='Imports':
                    import_index=index
                if "Domestic\nTotal" in cell:
                    domestic_index_2=index
                if cell=='Exports':
                    exports_index=index
                if cell=='Ending\nStocks':
                    ending_stock_index=index 
            break
    date=''
    #extracting data from cols
    n = 0
    geoant = ""
    for index,row in enumerate(data[header_index:]):
        new_row=[]
        if 'Beginning\nStocks' in row:
            date=row[date_index]
            if n == 0:
                n = 1
            elif n == 1:
                n = 2
            else:
                n = 3
            continue
        if (row[report_date_index]=='' and row[date_index] =='') or '1/ Data based on local marketing years except' in row[date_index]:
            continue
        if 'Selected Other' in row[date_index]:
            continue
        #se verifica cuando una fila contiene estos caracteres y segun el caso asigna el grupo
        if 'World' in row[date_index]:
            grupo = 'Resumen'
        if 'Major Exporters' in row[date_index]:
            grupo = 'Major Exporters'
        if 'Major Importers' in row[date_index]:
            grupo = 'Major Importers'
        new_row.append('XLS')
        new_row.append(archivo)
        new_row.append(wasde.replace(' ','').strip().upper())
        new_row.append(datos.upper()) 
        new_row.append(commoditie.upper()) 
        new_row.append(medida.upper())  
        new_row.append(n)
        new_row.append(date.split()[0])
        if len(date.split()) == 2:
            new_row.append(date.split()[1].upper())
        else:
            new_row.append('')            
        new_row.append(grupo.upper())
        if row[date_index].strip() == '' and row[report_date_index].strip() != '':
            row[date_index] = data[header_index:][index - 1][date_index]
        geo = replace_string(row[date_index].upper().strip())
        if geo == geoant:
            orden = 2
        else:
            if n == 1 or n == 2:
                orden = 2
            elif n == 3:
                orden = 1
        new_row.append(geo)  #add geographic
        new_row.append(orden)
        if  row[report_date_index].upper() == '':
            new_row.append(obtener_mes(mid(archivo,4,2)).upper())#Mes columnas vacias porque solo se usan en pag 19        
        else:
            new_row.append(row[report_date_index].upper())
        new_row.append(convert_value(row[begin_stock_index]))
        new_row.append(convert_value(row[production_index]))
        new_row.append(convert_value(row[import_index]))
        new_row.append('')
        new_row.append(convert_value(row[domestic_index_2]))
        new_row.append(convert_value(row[exports_index]))
        new_row.append(convert_value(row[ending_stock_index]))
        total_use = float(convert_value(row[domestic_index_2])) + float(convert_value(row[exports_index]))
        if total_use == 0:
            stock_to_use = 0
        else:            
            stock_to_use = (float(convert_value(row[ending_stock_index])) / total_use) * 100
        new_row.append(total_use)
        new_row.append(stock_to_use)        
        new_data.append(new_row)
        geoant = geo
    return new_data

def replace_string(cadena):
    for ch in ['1/','2/','3/','4/','5/','6/','7/','8/','9/','10/']:
        if ch in cadena:
            cadena=cadena.replace(ch,"")
    return cadena 

def convert_value(valor):
    if valor == '' or valor == 'NA':
        valor = 0
    return valor 

def mid(s, offset, amount):
    return s[offset:offset+amount]

def left(s, amount):
    return s[:amount]

def right(s, amount):
    return s[-amount:]

def obtener_mes(month):
    month_dictionary = {'01':'jan','02':'feb','03':'mar','04':'apr','05':'may','06':'jun',
    '07':'jul','08':'aug','09':'sep','10':'oct','11':'nov','12':'dec'}
    return month_dictionary.get(month)

def exist_fileurl(url):
    request = requests.get(url)
    if request.status_code == 200:
        return True
    else:
        return False
